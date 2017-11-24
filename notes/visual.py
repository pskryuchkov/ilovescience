from networkx.drawing.nx_agraph import graphviz_layout
from gensim.models import Word2Vec, LdaMulticore
from IPython.display import display, Markdown
from openpyxl import load_workbook
from matplotlib import cm
import networkx as nx
from numpy import *

import plotly.graph_objs as go
import plotly.offline as py
from plotly import tools

import pickle
import os

from warnings import warn
import csv


chart_count = 0

cite_table_head = \
"""
| {0}Author(s){1} | {0}Year{1} | {0}Ref. count{1} |
| --- | --- | --- |
""".format("<center>", "<center>")


articles_table_head = \
"""
| Topic title | Articles |
| --- | --- |
"""


def replics(model, target_word, topn=10):
    return {word: model.wv.vocab[word].count
                           for word in model.wv.vocab.keys()
                           if word.find(target_word) == 0}


def topics_normalize(model, raw_topics):
    norm_topics = []
    for word in raw_topics:
        r = replics(model, word)
        if r is not None:
            norm_topics.append(max(r, key=r.get))
        else:
            norm_topics.append(None)

    return norm_topics


def sigmoid(x):
    return 1 / (1 + math.exp(-x))


def get_lines(fn):
    return [line.strip() for line in open(fn, "r").readlines()]


def color_convert(cmap, pl_entries):
    h = 1.0/(pl_entries-1)
    pl_colorscale = []

    for k in range(pl_entries):
        C = list(map(uint8, array(cmap(k*h)[:3])*255))
        pl_colorscale.append([k*h, 'rgb'+str((C[0], C[1], C[2]))])

    return pl_colorscale


def draw_heatmap(volume, terms):
    corr_path = "../stat/frequency/{}/terms_similar.csv".format(volume)
    corr = []
    with open(corr_path, "rt") as f:
        reader = csv.reader(f)
        next(f)
        next(f)
        for line in reader:
            corr.append([float(x) for x in line[1:]])

    blues_cmap = cm.get_cmap('Blues')
    blues = color_convert(blues_cmap, 255)

    layout = go.Layout(
        xaxis = dict(
            ticktext=terms,
            tickvals=[i for i in range(len(corr))],
            tickangle=-45,
            linewidth=1,
            mirror=True
        ),
        yaxis = dict(
            ticktext=terms,
            tickvals=[i for i in range(len(corr))],
            tickangle=-45,
            linewidth=1,
            mirror=True
        ),
        width = 500,
        height = 500,
        title = "",
        margin=go.Margin(t=50)
        )

    trace = go.Heatmap(z = corr,
                       colorscale = blues,
                       showscale = False)
    data=[trace]
    py.iplot({'data': data, 'layout': layout}, show_link=False)


def add_edges(G, unique_pairs, terms, model, min_dist):
    w = []
    for pair in unique_pairs:
        word_one = terms[pair[0]]
        word_two = terms[pair[1]]
        dist = model.similarity(word_one, word_two)
        if dist > min_dist:
            w.append(edge_width)
        else:
            w.append(0)
        G.add_edge(word_one, word_two)
    return w


def draw_graph(terms, extented_terms, model):
    draw_method = "neato"
    min_dist1 = 0.3
    min_dist2 = 0.85
    font_small = 10
    font_large = 15

    terms_u = terms + extented_terms

    G = nx.Graph()
    G.add_nodes_from(terms_u)
    p = graphviz_layout(G, prog=draw_method)

    trace = go.Scatter(
        x = [p[word][0] for word in terms_u],
        y = [p[word][1] for word in terms_u],
        mode = 'text',
        name = 'None',
        text = ["<b>" + x + "</b>" if x in terms else x for x in terms_u],
        textposition = 'center',
        textfont = {
                    "size": [font_large if x in terms else font_small for x in terms_u]
                }
    )
    shapes = []
    unique_pairs = [[i, j] for j in range(len(terms_u)) for i in range(j)]

    for pair in unique_pairs:
            word_one = terms_u[pair[0]]
            word_two = terms_u[pair[1]]
            dist = model.similarity(word_one, word_two)
            if word_one in terms and word_two in terms:
                if dist > min_dist1:
                    shapes.append({
                        'type': 'line',
                        'x0': p[word_one][0],
                        'y0': p[word_one][1],
                        'x1': p[word_two][0],
                        'y1': p[word_two][1],
                        'line': {
                                'color': "red",
                                'width': 1.0
                                },
                        'layer' : 'below',
                        'opacity' : 0.65
                        })
            else:
                if dist > min_dist2:
                    shapes.append({
                        'type': 'line',
                        'x0': p[word_one][0],
                        'y0': p[word_one][1],
                        'x1': p[word_two][0],
                        'y1': p[word_two][1],
                        'line': {
                                'color': "lightskyblue",
                                'width': 1.0
                                },
                        'layer' : 'below',
                        'opacity' : 0.65
                        })

    layout = go.Layout(
        showlegend=False,
        shapes = shapes,
        xaxis=dict(
            showgrid=False,
            showticklabels = False,
            zeroline=False),
        yaxis=dict(
            showgrid=False,
            showticklabels = False,
            zeroline=False),
        width=700,
        height=700,
        hovermode = 'closest',
        title = ""
    )

    data = [trace]
    py.iplot({'data': data, 'layout': layout}, show_link=False)


def draw_scatter(volume):
    n_words = 50

    wb = load_workbook(filename = '../stat/frequency/{}/topics.xlsx'.format(volume))

    sheet_list = wb.get_sheet_names()
    sheet_list.remove('uncovered')

    data = []
    for name in sheet_list:
        ws = wb[name]

        if min(ws.max_row, ws.max_column) <= 1:
            warn("Empty sheet '{}'".format(name))
            continue

        table = [[line[0].value, line[1].value, line[2].value] for line in
                    zip(list(ws.columns)[0], list(ws.columns)[1], list(ws.columns)[2])]

        labels = [line[0] for line in table]
        x = [line[1] for line in table]
        y = [line[2] for line in table]

        data.append(dict(
                        type = 'scatter',
                        x = [line[1] for line in table][:n_words],
                        y = [line[2] for line in table][:n_words],
                        mode = 'markers',
                        hoverinfo = "text",
                        hoveron = "points",
                        name = name,
                        text = [line[0] for line in table][:n_words]
                    ))

    layout = dict(
            hovermode= 'closest',
            xaxis=dict(title="tf-idf"),
            yaxis=dict(title="frequency"),
            title = ""
            )

    py.iplot({'data': data, 'layout': layout}, validate=False, show_link=False)


def add_bar(fig, inp, n_topics, n_keys, n_row, n_col):
    global chart_count

    if chart_count >= n_topics:
        return

    keys = sorted([x for x in inp[:n_keys]], key=lambda x: x[1])

    labels = [x[1] for x in keys]
    values = [x[0] for x in keys]

    data = go.Bar(
            x=labels,
            y=values,
            orientation = 'h',
            name = "topic #{}".format(chart_count + 1),
            marker = dict(color='orange')
    )
    fig.append_trace(data, (chart_count // n_col) + 1, (chart_count % n_col) + 1)
    fig['layout']['xaxis{0}'.format(chart_count+1)].update(showticklabels = False)
    fig['layout']['yaxis{0}'.format(chart_count+1)].update(ticklen=3)
    fig['layout']['yaxis{0}'.format(chart_count+1)]['tickfont'].update(size=11,
                                                                       color="lightgrey")
    chart_count += 1


def draw_barchart(volume, n_keys = 10, n_row = 5, n_col = 3):
    fn = "../stat/lda/{}.keys.csv".format(volume)
    n_charts = n_row * n_col

    table = []
    val = []
    key = []
    with open(fn, "rt") as f:
        reader = csv.reader(f)
        next(f)
        for line in reader:
            if len(line) == 1:
                table.append([])
            else:
                table[-1].append([line[1], float(line[0])])

    n_topics = len(table)

    fig = tools.make_subplots(rows=n_row, cols=n_col, print_grid=False,
                              horizontal_spacing = 0.15, vertical_spacing=0.05,
                              subplot_titles=[x[0][0] for x in table[:n_charts]])

    for topic_words in table[:n_charts]:
        add_bar(fig, topic_words, n_topics, n_keys, n_row, n_col)

    fig['layout'].update(height=750, width=750, showlegend=False, title = "", margin=go.Margin(t=120))
    py.iplot(fig, show_link=False);


def cites(volume, n_cites=10, head=cite_table_head):
    markdown_table = head

    with open("../stat/references/{}.txt".format(volume), "r") as f:
        data = f.readlines()

    line_format = "|{0}{2}{1}|{0}{3}{1}|{0}{4}{1}|\n"
    for line in data[:n_cites]:
        authors, year, count = line.split(",")

        markdown_table += line_format.format("<center>", "</center>",
                                              ", ".join(map(lambda x: x.strip(), authors.split("&"))),
                                              year.strip(), count.strip())
    display(Markdown(markdown_table))


def top_chart(volume):
    data = open("../stat/frequency/{0}/terms_unique.csv".format(volume), "r").readlines()[2:]
    keywords = []
    for line in data:
        word, val, _ = line.split(",")
        keywords.append([word, float(val)])

    keywords_s = sorted(keywords, key = lambda x: x[1], reverse=True)

    data = [go.Bar(
                x=[x[0] for x in keywords_s],
                y=[x[1] for x in keywords_s]
            )]

    layout = go.Layout(title="")

    py.iplot({'data': data, 'layout': layout}, show_link = False)


def fn_pure(fn):
    return os.path.splitext(os.path.basename(fn))[0]


def lda_table(volume):
    fn = "../stat/lda/{}.keys.csv".format(volume)
    table = []
    with open(fn, "rt") as f:
        reader = csv.reader(f)
        next(f)
        for line in reader:
            if len(line) == 1:
                table.append([])
            else:
                table[-1].append([line[1], float(line[0])])
    return table


def topic_occur(text, topic, max_len=5000):
    if len(text) > max_len:
        return 0

    counter = 0
    for word in topic.keys():
        counter += text.count(word) * topic[word]
    return counter


def relevant_articles(volume, n_articles=1000, n_topic_articles=5):
    table = lda_table(volume)
    text_cache = pickle.load(open('../src/extra/{}.cache'.format(volume), 'rb'))

    topics_list = []
    topic_titles = []

    for x in table:
        topic = {}
        for y in x:
            topic[y[0]] = y[1]

        topic_titles.append(x[0][0])
        topics_list.append(topic)

    text_indexes = text_cache.keys()

    score_list = []
    for topic in topics_list:
        occur = []

        for idx in text_indexes[:n_articles]:
            text = text_cache[idx][0].split()

            val = 1.0 * topic_occur(text, topic)
            if val > 0:
                occur.append([fn_pure(idx), val])

        score_list.append(occur)

    sorted_score = [sorted(x, key=lambda z: z[1], reverse=True) for x in score_list]

    relevant = []
    for z in sorted_score:
        sub = []

        for w in z[:n_topic_articles]:
            sub.append(w[0])

        relevant.append(sub)

    return topic_titles, relevant


def lda_print(topic_names, articles_table, head=articles_table_head):
    markdown_table = head

    arxiv_link = "[{0}](https://arxiv.org/pdf/{0}.pdf)" + "&nbsp;" * 4

    for j, topic in enumerate(topic_names):

        ref_s = ""
        for article_id in articles_table[j]:
            ref_s += str(arxiv_link.format(article_id))

        markdown_table += "| {0} | {1} |\n".format(topic, ref_s)

    display(Markdown(markdown_table))


def lda_articles(volume):
    titles, articles = relevant_articles(volume)
    lda_print(titles, articles)